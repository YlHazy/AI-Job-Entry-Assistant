"""Detect and cache mappings for arbitrary Excel job board templates."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.config import BAILIAN_BASE_URL, BAILIAN_MODEL, TEMPLATE_CACHE_PATH, get_bailian_api_key
from src.prompt_loader import load_prompt
from src.schemas import ExcelTemplateMapping

CANONICAL_FIELD_ALIASES = {
    "company": ["公司", "企业", "company"],
    "job_title": ["岗位名称", "职位名称", "岗位", "job title"],
    "role_category": ["岗位类别", "岗位类型", "role category"],
    "direction": ["方向", "匹配方向", "match direction"],
    "location": ["城市/办公形式", "工作地点", "地点", "location"],
    "jd_url": ["JD链接", "来源链接", "岗位链接", "url"],
    "source_platform": ["来源渠道", "来源平台", "source"],
    "applied_date": ["投递日期", "录入日期"],
    "resume_customized": ["是否定制", "是否建议单独改简历"],
    "keyword_hit_level": ["关键词命中度", "关键词命中"],
    "priority": ["优先级"],
    "status": ["当前状态", "状态"],
    "risk_note": ["风险/挂点", "挂点分析", "主要差距提示"],
    "summary_note": ["复盘总结", "备注", "投递备注"],
}


class ExcelTemplateAdapter:
    """Resolve a workbook-specific template mapping."""

    def __init__(
        self,
        cache_path: Path = TEMPLATE_CACHE_PATH,
        api_key: str | None = None,
        base_url: str = BAILIAN_BASE_URL,
        model: str = BAILIAN_MODEL,
        client: OpenAI | None = None,
    ) -> None:
        self.cache_path = cache_path
        self.api_key = (api_key or get_bailian_api_key()).strip()
        self.base_url = base_url
        self.model = model
        self.client = client or (OpenAI(api_key=self.api_key, base_url=self.base_url) if self.api_key else None)

    def resolve(self, workbook_path: str | Path, requested_sheet: str = "") -> ExcelTemplateMapping:
        workbook_path = Path(workbook_path).expanduser().resolve()
        workbook_key = compute_workbook_key(workbook_path)
        cached = self._load_cached_mapping(workbook_key, requested_sheet)
        if cached:
            return cached

        workbook = load_workbook(workbook_path)
        try:
            mapping = self._detect_mapping(workbook, workbook_key, requested_sheet)
        finally:
            workbook.close()
        self._save_cached_mapping(mapping)
        return mapping

    def _detect_mapping(self, workbook, workbook_key: str, requested_sheet: str) -> ExcelTemplateMapping:
        candidates: list[ExcelTemplateMapping] = []
        for worksheet in workbook.worksheets:
            if requested_sheet and worksheet.title != requested_sheet:
                continue
            rule_mapping = detect_with_rules(worksheet, workbook_key)
            candidates.append(rule_mapping)

        if not candidates:
            raise ValueError("No worksheet is available for template detection.")

        best = max(candidates, key=lambda item: item.confidence)
        if best.confidence >= 0.75 or not self.client:
            return best

        llm_mapping = self._detect_with_llm(workbook, workbook_key, requested_sheet, best)
        return llm_mapping if llm_mapping.confidence >= best.confidence else best

    def _detect_with_llm(
        self,
        workbook,
        workbook_key: str,
        requested_sheet: str,
        fallback: ExcelTemplateMapping,
    ) -> ExcelTemplateMapping:
        prompt = load_prompt("excel_template_mapping.md")
        workbook_snapshot = build_workbook_snapshot(workbook, requested_sheet=requested_sheet)
        user_prompt = prompt.replace("{{workbook_snapshot}}", workbook_snapshot)
        response = self.client.chat.completions.create(
            model=self.model,
            temperature=0,
            messages=[
                {"role": "system", "content": "Return exactly one JSON object and nothing else."},
                {"role": "user", "content": user_prompt},
            ],
        )
        content = response.choices[0].message.content or ""
        payload = parse_json_payload(content)

        sheet_name = str(payload.get("target_sheet", fallback.target_sheet)).strip() or fallback.target_sheet
        return ExcelTemplateMapping(
            workbook_key=workbook_key,
            target_sheet=sheet_name,
            header_row=max(1, int(payload.get("header_row", fallback.header_row))),
            data_start_row=max(1, int(payload.get("data_start_row", fallback.data_start_row))),
            key_columns={k: int(v) for k, v in dict(payload.get("key_columns", fallback.key_columns)).items() if int(v) > 0},
            confidence=float(payload.get("confidence", fallback.confidence)),
            warnings=[str(item) for item in payload.get("warnings", fallback.warnings)],
            detection_mode="llm",
        )

    def _load_cached_mapping(self, workbook_key: str, requested_sheet: str) -> ExcelTemplateMapping | None:
        cache = load_template_cache(self.cache_path)
        payload = cache.get(workbook_key)
        if not payload:
            return None
        mapping = ExcelTemplateMapping.model_validate(payload)
        if requested_sheet and mapping.target_sheet != requested_sheet:
            return None
        return mapping

    def _save_cached_mapping(self, mapping: ExcelTemplateMapping) -> None:
        cache = load_template_cache(self.cache_path)
        cache[mapping.workbook_key] = mapping.model_dump()
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        self.cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def detect_with_rules(worksheet: Worksheet, workbook_key: str) -> ExcelTemplateMapping:
    best_row = 1
    best_score = -1.0
    best_columns: dict[str, int] = {}
    flattened_aliases = flatten_aliases()

    for row_index in range(1, min(worksheet.max_row, 20) + 1):
        headers = [normalize_text(worksheet.cell(row_index, column).value) for column in range(1, worksheet.max_column + 1)]
        score = sum(1 for value in headers if value in flattened_aliases)
        columns = resolve_key_columns(headers)
        if score > best_score:
            best_score = float(score)
            best_row = row_index
            best_columns = columns

    confidence = min(0.95, 0.2 + 0.08 * len(best_columns))
    warnings = []
    if best_row > 3:
        warnings.append("Header row is not at the top of the sheet.")
    if worksheet.max_row - best_row > 50:
        warnings.append("The sheet looks like a preformatted board with a reserved data range.")

    return ExcelTemplateMapping(
        workbook_key=workbook_key,
        target_sheet=worksheet.title,
        header_row=best_row,
        data_start_row=best_row + 1,
        key_columns=best_columns,
        confidence=confidence,
        warnings=warnings,
        detection_mode="rules",
    )


def resolve_key_columns(headers: list[str]) -> dict[str, int]:
    resolved: dict[str, int] = {}
    for canonical, aliases in CANONICAL_FIELD_ALIASES.items():
        for index, header in enumerate(headers, start=1):
            if header and header in {normalize_text(alias) for alias in aliases}:
                resolved[canonical] = index
                break
    return resolved


def flatten_aliases() -> set[str]:
    values: set[str] = set()
    for aliases in CANONICAL_FIELD_ALIASES.values():
        values.update(normalize_text(alias) for alias in aliases)
    return values


def normalize_text(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def compute_workbook_key(path: Path) -> str:
    raw = f"{path.resolve()}::{path.stat().st_size}::{int(path.stat().st_mtime)}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def load_template_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def build_workbook_snapshot(workbook, requested_sheet: str = "") -> str:
    parts: list[str] = []
    for worksheet in workbook.worksheets:
        if requested_sheet and worksheet.title != requested_sheet:
            continue
        parts.append(f"[Sheet] {worksheet.title}")
        for row_index in range(1, min(worksheet.max_row, 12) + 1):
            row_values = [worksheet.cell(row_index, col).value for col in range(1, min(worksheet.max_column, 12) + 1)]
            parts.append(f"row {row_index}: {row_values}")
    return "\n".join(parts)


def parse_json_payload(content: str) -> dict[str, Any]:
    cleaned = content.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.removeprefix("```json").removeprefix("```JSON").removeprefix("```").strip()
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3].strip()
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1:
        raise ValueError("Template mapping response does not contain JSON.")
    return json.loads(cleaned[start : end + 1])
