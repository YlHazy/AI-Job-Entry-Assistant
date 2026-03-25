"""Excel writer with adaptive template detection."""

from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from src.excel_template_adapter import ExcelTemplateAdapter
from src.path_utils import normalize_sheet_name, normalize_user_text
from src.schemas import ExcelTemplateMapping, JobRecord

DEFAULT_HEADERS = [
    "公司",
    "岗位名称",
    "岗位类别",
    "方向",
    "城市/办公形式",
    "JD链接",
    "来源渠道",
    "投递日期",
    "是否定制",
    "关键词命中度",
    "优先级",
    "当前状态",
    "风险/挂点",
    "复盘总结",
]


class ExcelWriteError(RuntimeError):
    """Raised when the app cannot safely write to the target workbook."""

    def __init__(self, message: str, reasons: list[str] | None = None) -> None:
        super().__init__(message)
        self.reasons = reasons or []


def append_job_record(
    excel_path: str | Path,
    record: JobRecord,
    sheet_name: str = "",
    adapter: ExcelTemplateAdapter | None = None,
) -> Path:
    """Write a job record into an existing board-friendly workbook."""

    normalized_path = normalize_user_text(str(excel_path))
    normalized_sheet = normalize_sheet_name(sheet_name)
    if not normalized_path:
        raise ExcelWriteError("未提供 Excel 文件路径。", ["请在设置页填写 Excel 路径。"])

    path = Path(normalized_path).expanduser().resolve()
    workbook = load_or_create_workbook(path, normalized_sheet or "JobBoard")
    try:
        if path.exists():
            mapping = (adapter or ExcelTemplateAdapter()).resolve(path, requested_sheet=normalized_sheet)
        else:
            worksheet = workbook[normalized_sheet or "JobBoard"]
            ensure_headers(worksheet)
            mapping = ExcelTemplateMapping(
                workbook_key="new-workbook",
                target_sheet=normalized_sheet or "JobBoard",
                header_row=1,
                data_start_row=2,
                key_columns={field: index + 1 for index, field in enumerate(default_column_order())},
                confidence=1.0,
                warnings=[],
                detection_mode="default",
            )

        worksheet = workbook[mapping.target_sheet]
        target_row = find_next_available_row(worksheet, mapping)
        row_values = build_row_values(record)

        for field, value in row_values.items():
            column_index = mapping.key_columns.get(field)
            if column_index is None:
                continue
            worksheet.cell(row=target_row, column=column_index, value=value)

        apply_row_presentation(worksheet, target_row, mapping)
        try:
            workbook.save(path)
        except PermissionError as exc:
            raise ExcelWriteError(
                "无法写入 Excel 文件。",
                [
                    "文件可能正被 Excel 或 WPS 打开，请先关闭后重试。",
                    "当前路径可能没有写权限。",
                    "如果路径在网盘或同步目录中，请确认该文件没有被占用。",
                ],
            ) from exc
        except OSError as exc:
            raise ExcelWriteError(
                f"写入 Excel 时发生系统错误：{exc}",
                [
                    "请确认文件路径存在且格式正确。",
                    "如果路径包含中文或被引号包裹，应用现在支持自动清理，但仍建议再核对一次。",
                ],
            ) from exc
        return path
    finally:
        workbook.close()


def load_or_create_workbook(path: Path, sheet_name: str) -> Workbook:
    if path.exists():
        workbook = load_workbook(path)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        return workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    return workbook


def ensure_headers(worksheet: Worksheet) -> None:
    if worksheet.max_row >= 1 and any(cell.value for cell in worksheet[1]):
        return
    for column_index, header in enumerate(DEFAULT_HEADERS, start=1):
        worksheet.cell(row=1, column=column_index, value=header)


def default_column_order() -> list[str]:
    return [
        "company",
        "job_title",
        "role_category",
        "direction",
        "location",
        "jd_url",
        "source_platform",
        "applied_date",
        "resume_customized",
        "keyword_hit_level",
        "priority",
        "status",
        "risk_note",
        "summary_note",
    ]


def find_next_available_row(worksheet: Worksheet, mapping: ExcelTemplateMapping) -> int:
    required_fields = [field for field in ["company", "job_title"] if field in mapping.key_columns]
    if not required_fields:
        return max(mapping.data_start_row, worksheet.max_row + 1)

    for row_index in range(mapping.data_start_row, worksheet.max_row + 1):
        if all(not cell_has_meaningful_value(worksheet.cell(row_index, mapping.key_columns[field]).value) for field in required_fields):
            return row_index
    return worksheet.max_row + 1


def build_row_values(record: JobRecord) -> dict[str, Any]:
    return {
        "company": record.company,
        "job_title": record.job_title,
        "role_category": record.role_category,
        "direction": record.match_direction,
        "location": record.location,
        "jd_url": record.source_url,
        "source_platform": record.source_platform,
        "applied_date": datetime.now().strftime("%Y-%m-%d"),
        "resume_customized": record.recommend_resume_customization,
        "keyword_hit_level": infer_keyword_hit_level(record),
        "priority": normalize_priority_for_board(record.priority),
        "status": "待投递",
        "risk_note": record.main_gap or record.notes,
        "summary_note": record.short_note or record.notes,
    }


def infer_keyword_hit_level(record: JobRecord) -> str:
    count = len(record.keywords)
    if count >= 6:
        return "高"
    if count >= 3:
        return "中"
    if count >= 1:
        return "低"
    return ""


def normalize_priority_for_board(priority: str) -> str:
    mapping = {"高": "P1", "中": "P2", "低": "P3", "": ""}
    return mapping.get(priority, priority)


def cell_has_meaningful_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.startswith("="):
        return False
    return str(value).strip() != ""


def apply_row_presentation(worksheet: Worksheet, target_row: int, mapping: ExcelTemplateMapping) -> None:
    template_row = mapping.data_start_row
    base_height = worksheet.row_dimensions[template_row].height or 22
    worksheet.row_dimensions[target_row].height = min(base_height, 24)

    for column_index in mapping.key_columns.values():
        cell = worksheet.cell(target_row, column_index)
        template_cell = worksheet.cell(template_row, column_index)
        if template_cell.has_style:
            cell._style = copy(template_cell._style)
            current_alignment = copy(cell.alignment)
            current_alignment.wrap_text = False
            current_alignment.vertical = "center"
            cell.alignment = current_alignment
        else:
            cell.font = Font(name="Microsoft YaHei UI", size=10.5)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
