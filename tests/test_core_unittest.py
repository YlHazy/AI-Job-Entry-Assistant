import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.agent import JobIntakeAgent
from src.classifier import enrich_job_record
from src.deduper import looks_like_duplicate
from src.excel_template_adapter import ExcelTemplateAdapter
from src.excel_writer import ExcelWriteError, append_job_record
from src.history_store import HistoryStore
from src.parser import parse_job_text
from src.schemas import JobRecord
from src.web_fetcher import WebFetchResult


class ParserTests(unittest.TestCase):
    def test_parse_job_text_extracts_core_fields(self) -> None:
        raw_text = """
        公司：星河智能科技
        岗位名称：AI Agent 开发实习生
        工作地点：上海
        岗位职责：
        负责基于 Python、RAG、LLM 和工作流搭建智能助手。
        任职要求：
        本科及以上，熟悉 Prompt Engineering、OpenAI API、LangChain。
        """

        record = parse_job_text(raw_text, source_platform="实习僧", source_url="https://example.com/job")

        self.assertEqual(record.company, "星河智能科技")
        self.assertEqual(record.job_title, "AI Agent 开发实习生")
        self.assertEqual(record.location, "上海")
        self.assertTrue(record.is_internship)
        self.assertEqual(record.education_requirement, "本科及以上")
        self.assertIn("Python", record.skills)
        self.assertIn("LLM", record.keywords)


class ClassifierTests(unittest.TestCase):
    def test_enrich_job_record_marks_agent_role_high_priority(self) -> None:
        record = JobRecord(
            job_title="AI Agent 开发实习生",
            jd_summary="负责 Agent workflow、RAG、LLM 应用开发和 Python 后端能力建设。",
            keywords=["LLM", "RAG", "工作流"],
            skills=["Python", "OpenAI", "LangChain"],
            raw_text="需要负责 Agent、RAG、Prompt、Python 开发。",
        )

        enriched = enrich_job_record(record)

        self.assertEqual(enriched.role_category, "开发")
        self.assertEqual(enriched.match_direction, "AI Agent 开发")
        self.assertEqual(enriched.priority, "高")
        self.assertEqual(enriched.recommend_resume_customization, "是")


class ExcelWriterTests(unittest.TestCase):
    def test_append_job_record_creates_workbook_and_sheet(self) -> None:
        tmp_dir = Path(__file__).resolve().parent / "_tmp"
        tmp_dir.mkdir(exist_ok=True)
        excel_path = tmp_dir / "jobs.xlsx"
        if excel_path.exists():
            excel_path.unlink()

        record = JobRecord(
            company="星河智能科技",
            job_title="AI Agent 开发实习生",
            role_category="开发",
            match_direction="AI Agent 开发",
            priority="高",
            recommend_resume_customization="是",
            location="上海",
            source_platform="实习僧",
        )

        append_job_record(excel_path, record)

        workbook = load_workbook(excel_path)
        worksheet = workbook["JobBoard"]
        self.assertEqual(worksheet["A2"].value, "星河智能科技")
        self.assertEqual(worksheet["B2"].value, "AI Agent 开发实习生")
        self.assertEqual(worksheet["C2"].value, "开发")
        workbook.close()

    def test_append_job_record_requires_path(self) -> None:
        with self.assertRaises(ExcelWriteError):
            append_job_record("", JobRecord(company="星河智能科技"))

    def test_append_job_record_uses_existing_board_template(self) -> None:
        tmp_dir = Path(__file__).resolve().parent / "_tmp"
        tmp_dir.mkdir(exist_ok=True)
        excel_path = tmp_dir / "template_board.xlsx"
        if excel_path.exists():
            excel_path.unlink()

        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JobBoard"
        worksheet["A1"] = "说明"
        headers = [
            "编号",
            "公司",
            "岗位名称",
            "岗位类别",
            "方向",
            "城市/办公形式",
            "JD链接",
            "来源渠道",
            "投递日期",
            "截止日期",
            "简历版本",
            "是否定制",
            "关键词命中度",
            "优先级",
            "当前状态",
            "最新进展日期",
            "距上次进展(天)",
            "下一步动作",
            "下一步截止",
            "面试轮次",
            "笔试情况",
            "联系人/内推",
            "匹配度(1-5)",
            "风险/挂点",
            "复盘总结",
        ]
        for idx, header in enumerate(headers, start=1):
            worksheet.cell(row=7, column=idx, value=header)
        worksheet["A8"] = '=IF(B8="","",ROW()-7)'
        worksheet["Q8"] = '=IF(P8="","",TODAY()-P8)'
        workbook.save(excel_path)
        workbook.close()

        record = JobRecord(
            company="星河智能科技",
            job_title="AI Agent 开发实习生",
            role_category="开发",
            match_direction="AI Agent 开发",
            priority="高",
            recommend_resume_customization="是",
            location="上海",
            source_platform="官网",
            source_url="https://example.com/job",
            main_gap="需要强化 Agent 项目经历",
            short_note="值得优先投递",
            keywords=["LLM", "RAG", "Agent", "Python"],
        )

        adapter = ExcelTemplateAdapter()
        append_job_record(excel_path, record, adapter=adapter)

        workbook = load_workbook(excel_path)
        worksheet = workbook["JobBoard"]
        self.assertEqual(worksheet["B8"].value, "星河智能科技")
        self.assertEqual(worksheet["C8"].value, "AI Agent 开发实习生")
        self.assertEqual(worksheet["D8"].value, "开发")
        self.assertEqual(worksheet["E8"].value, "AI Agent 开发")
        self.assertEqual(worksheet["F8"].value, "上海")
        self.assertEqual(worksheet["G8"].value, "https://example.com/job")
        self.assertEqual(worksheet["H8"].value, "官网")
        self.assertEqual(worksheet["L8"].value, "是")
        self.assertEqual(worksheet["N8"].value, "P1")
        self.assertEqual(worksheet["O8"].value, "待投递")
        self.assertEqual(worksheet["X8"].value, "需要强化 Agent 项目经历")
        self.assertEqual(worksheet["Y8"].value, "值得优先投递")
        workbook.close()


class TemplateAdapterTests(unittest.TestCase):
    def test_template_adapter_detects_existing_board_shape(self) -> None:
        tmp_dir = Path(__file__).resolve().parent / "_tmp"
        tmp_dir.mkdir(exist_ok=True)
        excel_path = tmp_dir / "adapter_board.xlsx"
        if excel_path.exists():
            excel_path.unlink()

        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JobBoard"
        worksheet["A1"] = "说明"
        headers = ["编号", "公司", "岗位名称", "岗位类别", "方向", "城市/办公形式", "JD链接", "来源渠道", "投递日期", "是否定制", "关键词命中度", "优先级", "当前状态", "风险/挂点", "复盘总结"]
        for idx, header in enumerate(headers, start=1):
            worksheet.cell(row=7, column=idx, value=header)
        workbook.save(excel_path)
        workbook.close()

        adapter = ExcelTemplateAdapter(cache_path=tmp_dir / "template_cache.json", client=None)
        mapping = adapter.resolve(excel_path, requested_sheet="JobBoard")
        self.assertEqual(mapping.target_sheet, "JobBoard")
        self.assertEqual(mapping.header_row, 7)
        self.assertEqual(mapping.data_start_row, 8)
        self.assertEqual(mapping.key_columns["company"], 2)
        self.assertGreaterEqual(mapping.confidence, 0.75)


class HistoryStoreTests(unittest.TestCase):
    def test_history_store_inserts_rows(self) -> None:
        tmp_dir = Path(__file__).resolve().parent / "_tmp"
        tmp_dir.mkdir(exist_ok=True)
        db_path = tmp_dir / "history.db"
        if db_path.exists():
            db_path.unlink()

        store = HistoryStore(db_path=db_path)
        store.insert(
            JobRecord(
                company="星河智能科技",
                job_title="AI Agent 开发实习生",
                role_category="开发",
                match_direction="AI Agent 开发",
                priority="高",
            )
        )
        self.assertEqual(store.count(), 1)
        recent = store.recent(limit=5)
        self.assertEqual(len(recent), 1)
        self.assertEqual(recent[0]["company"], "星河智能科技")
        search_rows = store.search(query="星河", limit=5)
        self.assertEqual(len(search_rows), 1)
        self.assertEqual(search_rows[0]["job_title"], "AI Agent 开发实习生")


class DeduperTests(unittest.TestCase):
    def test_deduper_detects_same_url(self) -> None:
        record = JobRecord(company="星河智能科技", job_title="AI Agent 开发实习生", source_url="https://example.com/job")
        result = looks_like_duplicate(
            record,
            [{"id": 3, "company": "星河智能科技", "job_title": "AI Agent 开发实习生", "source_url": "https://example.com/job"}],
        )
        self.assertTrue(result.is_duplicate)
        self.assertEqual(result.reason, "来源链接重复")

    def test_deduper_detects_similar_company_and_title(self) -> None:
        record = JobRecord(company="星河智能科技", job_title="AI Agent 开发实习")
        result = looks_like_duplicate(
            record,
            [{"id": 4, "company": "星河智能科技", "job_title": "AI Agent 开发实习生", "source_url": ""}],
        )
        self.assertTrue(result.is_duplicate)
        self.assertEqual(result.reason, "公司和岗位名称高度相似")


class AgentTests(unittest.TestCase):
    def test_job_intake_agent_returns_trace_and_record(self) -> None:
        agent = JobIntakeAgent()
        raw_text = """
        公司：星河智能科技
        岗位名称：AI Agent 开发实习生
        工作地点：上海
        负责基于 Python、RAG、LLM、Prompt 和工作流搭建智能助手。
        本科及以上，熟悉 API 集成和 Agent workflow。
        """

        result = agent.run(raw_text, source_platform="实习僧", source_url="https://example.com/job", prefer_llm=False)

        self.assertEqual(result.record.company, "星河智能科技")
        self.assertEqual(result.record.role_category, "开发")
        self.assertEqual(result.record.match_direction, "AI Agent 开发")
        self.assertEqual(result.record.priority, "高")
        self.assertEqual(len(result.steps), 5)
        self.assertEqual(result.steps[0].name, "Extract core fields")
        self.assertEqual(result.steps[-1].name, "Prepare human review")

    def test_job_intake_agent_can_run_from_url(self) -> None:
        agent = JobIntakeAgent()
        agent.fetcher = lambda url: WebFetchResult(
            url=url,
            title="岗位详情页",
            text=(
                "公司：星河智能科技\n"
                "岗位名称：AI Agent 开发实习生\n"
                "工作地点：上海\n"
                "负责 Python、RAG、LLM 和 Agent workflow。\n"
                "本科及以上。"
            ),
        )

        result = agent.run_from_url("https://example.com/job", source_platform="官网", prefer_llm=False)

        self.assertEqual(result.record.company, "星河智能科技")
        self.assertEqual(result.steps[0].name, "Fetch job page")
        self.assertEqual(result.steps[1].name, "Extract core fields")


if __name__ == "__main__":
    unittest.main()
